# FILE: core/finance/ohlcv.py
# @core-candidate: OHLCVDownloader, 2026-04, OHLCV 데이터 수집 및 캐싱 엔진

import os
import sys
import time
import pickle
import logging
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
import pandas as pd
import yfinance as yf

from datetime import datetime

# 로거 설정 (기존 핸들러 설정을 따르되, core 레벨에서는 설정을 강제하지 않음)
log = logging.getLogger(__name__)

class OHLCVDownloader:
    """
    yfinance를 사용하여 OHLCV(Open, High, Low, Close, Volume) 데이터를 
    다운로드하고 로컬 디렉토리에 캐싱하는 클래스입니다.
    """

    def __init__(
        self, 
        cache_dir: str | Path = "ohlcv_cache", 
        period: str = "3y", 
        interval: str = "1d",
        workers: int = 8
    ):
        """
        Args:
            cache_dir (str | Path): 데이터가 저장될 폴더 경로
            period (str): 다운로드 기간 (예: '1y', '3y', 'max')
            interval (str): 데이터 간격 (예: '1d', '1h', '5m')
            workers (int): 병렬 다운로드 스레드 수
        """
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self.period = period
        self.interval = interval
        self.workers = workers

    def load_tickers_from_excel(self, file_path: str | Path) -> List[Dict[str, str]]:
        """
        Excel 파일에서 티커 리스트를 로드합니다.
        (첫 번째 열: 종목코드, 두 번째 열: 종목명 순서 기대)
        
        Args:
            file_path (str | Path): Excel 파일 경로
            
        Returns:
            List[Dict[str, str]]: [{'code': '005930', 'name': '삼성전자'}, ...]
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel 파일을 찾을 수 없습니다: {path}")

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        tickers = []
        for r in range(2, ws.max_row + 1):
            code = ws.cell(r, 1).value
            name = ws.cell(r, 2).value
            if code:
                # 한국 주식 종목코드는 6자리이므로 zfill(6) 처리
                tickers.append({"code": str(code).zfill(6), "name": str(name)})
        
        log.info(f"티커 리스트 로드 완료: {len(tickers)}종목 ({path.name})")
        return tickers

    def download_one(self, ticker_info: Dict[str, str], force: bool = False, auto_update: bool = True) -> Tuple[str, bool, str]:
        """
        단일 종목을 다운로드하고 캐시(pickle)로 저장합니다.
        
        Args:
            ticker_info (Dict): {'code': '...', 'name': '...'} 형태
            force (bool): 이미 파일이 있어도 강제 다운로드 여부
            auto_update (bool): 파일이 오늘 수정된 것이 아니라면 자동으로 업데이트할지 여부
            
        Returns:
            Tuple[str, bool, str]: (종목코드, 성공여부, 메시지)
        """
        code = ticker_info["code"]
        name = ticker_info["name"]
        
        # 한국 주식인 경우 .KS 접미사 사용 (추후 확장 가능)
        yf_symbol = f"{code}.KS"
        out_path = self.cache_dir / f"{code}.pkl"

        if not force and out_path.exists():
            if auto_update:
                # 파일 수정 날짜 확인
                mtime = datetime.fromtimestamp(out_path.stat().st_mtime).date()
                today = datetime.now().date()
                if mtime == today:
                    return code, True, f"{name} — 오늘 이미 업데이트됨, 스킵"
                else:
                    log.info(f"{name} — 데이터가 최신이 아님 (마지막 수정: {mtime}). 업데이트를 시작합니다.")
            else:
                return code, True, f"{name} — 캐시 존재, 스킵"

        try:
            # auto_adjust=True: 수정주가 반영
            raw = yf.download(
                yf_symbol, 
                period=self.period, 
                interval=self.interval,
                auto_adjust=True, 
                progress=False
            )
            
            if raw.empty:
                return code, False, f"{name} — 데이터가 비어 있음"

            # yfinance 다중 인덱스 또는 중복 컬럼 처리
            if isinstance(raw.columns, pd.MultiIndex):
                raw.columns = raw.columns.get_level_values(0)
            
            raw.columns = raw.columns.str.lower()
            raw = raw.loc[:, ~raw.columns.duplicated(keep="first")]

            if len(raw) < 5:  # 최소 데이터 검증
                return code, False, f"{name} — 데이터 부족 ({len(raw)}행)"

            # pkl로 저장
            with open(out_path, "wb") as f:
                pickle.dump(raw, f)

            return code, True, f"{name} — {len(raw)}일 저장 완료"

        except Exception as e:
            return code, False, f"{name} — 오류: {str(e)}"

    def download_all(self, tickers: List[Dict[str, str]], force: bool = False):
        """
        제공된 티커 리스트 전체를 병렬로 다운로드합니다.
        """
        total = len(tickers)
        success = 0
        fail = 0

        log.info(f"다운로드 시작: {total}종목 / {self.workers}스레드 병렬")

        with ThreadPoolExecutor(max_workers=self.workers) as executor:
            futures = {executor.submit(self.download_one, t, force): t for t in tickers}
            for i, future in enumerate(as_completed(futures), 1):
                code, ok, msg = future.result()
                if ok:
                    success += 1
                    log.info(f"[{i:3}/{total}] ✓ {msg}")
                else:
                    fail += 1
                    log.warning(f"[{i:3}/{total}] ✗ {msg}")

                # yfinance API 속도 제한(Rate Limit) 완화
                time.sleep(0.02)

        log.info(f"다운로드 완료: 성공 {success} / 실패 {fail} / 전체 {total}")
        log.info(f"캐시 위치: {self.cache_dir.absolute()}")

if __name__ == "__main__":
    # 간단한 테스트용 (core 모듈 단독 실행 시)
    logging.basicConfig(level=logging.INFO)
    downloader = OHLCVDownloader(cache_dir="test_cache", period="1y")
    # 예: 삼성전자(005930) 테스트
    res = downloader.download_one({"code": "005930", "name": "삼성전자"}, force=True)
    print(res)
